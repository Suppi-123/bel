import streamlit as st
import PyPDF2
import pandas as pd
import re
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import tempfile
from app.database.models import *
from datetime import datetime
from pony.orm import db_session

def clean_text(text):
    # Remove multiple spaces and clean up the text
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

def create_excel_file(document_details, operations_df, raw_materials_df=None):
    wb = Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Create Document Details sheet
    ws1 = wb.active
    ws1.title = "Document Details"
    
    # Add headers with styling
    ws1['A1'] = "Field"
    ws1['B1'] = "Value"
    ws1['A1'].fill = header_fill
    ws1['B1'].fill = header_fill
    ws1['A1'].font = header_font
    ws1['B1'].font = header_font
    
    # Add document details
    row = 2
    for key, value in document_details.items():
        ws1[f'A{row}'] = key
        ws1[f'B{row}'] = value
        row += 1
    
    # Adjust column widths
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 40
    
    # Create Operations sheet
    ws2 = wb.create_sheet("Operations")
    
    # Write operations data
    # Write headers
    for idx, col in enumerate(operations_df.columns, 1):
        cell = ws2.cell(row=1, column=idx)
        cell.value = col
        cell.fill = header_fill
        cell.font = header_font
    
    # Write data
    for r_idx, row in enumerate(operations_df.values, 2):
        for c_idx, value in enumerate(row, 1):
            ws2.cell(row=r_idx, column=c_idx, value=value)
    
    # Adjust column widths
    for column in ws2.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws2.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Create Raw Materials sheet if data exists
    if raw_materials_df is not None:
        ws3 = wb.create_sheet("Raw Materials")
        
        # Write headers
        headers = ["Sl.No", "Child Part No", "Description", "Qty Per Set", "UoM", "Total Qty"]
        for idx, header in enumerate(headers, 1):
            cell = ws3.cell(row=1, column=idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        # Write raw materials data
        for r_idx, row in enumerate(raw_materials_df.values, 2):
            for c_idx, value in enumerate(row, 1):
                ws3.cell(row=r_idx, column=c_idx, value=value)
        
        # Adjust column widths
        for column in ws3.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws3.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Save to temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        return tmp.name

def extract_oarc_details(pdf_content):
    # Read PDF
    pdf_reader = PyPDF2.PdfReader(pdf_content)
    text = ""
    for page in pdf_reader.pages:
        text += page.extract_text() + "\n"
    
    # Initialize dictionary to store extracted data
    data = {
        "Project Name": "",
        "Sale Order": "",
        "Part No": "",
        "Part Desc": "",
        "Required Qty": "",
        "Plant": "",
        "WBS": "",
        "Rtg Seq No": "",
        "Sequence No": "",
        "Launched Qty": "",
        "Prod Order No": "",
        "Operations": [],
        "Document Verification": {},
        "Raw Materials": []
    }
    
    # Extract header information using specific patterns
    # Project Name and Part No
    project_match = re.search(r"Project Name\s*:([^:]+)Part No\s*:([^W]+)WBS\s*:\s*([^\n]+)", text)
    if project_match:
        data["Project Name"] = project_match.group(1).strip()
        data["Part No"] = project_match.group(2).strip()
        data["WBS"] = project_match.group(3).strip()
    
    # Sale order and Part Desc
    sale_match = re.search(r"Sale order\s*:([^:]+)Part Desc\s*:([^T]+)", text)
    if sale_match:
        data["Sale Order"] = sale_match.group(1).strip()
        data["Part Desc"] = sale_match.group(2).strip()
    
    # Plant and sequence numbers
    plant_match = re.search(r"Plant\s*:([^R]+)Rtg Seq No\s*:([^S]+)Sequence No\s*:([^\n]+)", text)
    if plant_match:
        data["Plant"] = plant_match.group(1).strip()
        data["Rtg Seq No"] = plant_match.group(2).strip()
        data["Sequence No"] = plant_match.group(3).strip()
    
    # Required Qty, Launched Qty, and Prod Order No
    qty_match = re.search(r"Required Qty\s*:([^L]+)Launched Qty\s*:([^P]+)Prod Order No\s*:([^\n]+)", text)
    if qty_match:
        data["Required Qty"] = qty_match.group(1).strip()
        data["Launched Qty"] = qty_match.group(2).strip()
        data["Prod Order No"] = qty_match.group(3).strip()
    
    # Extract operations
    lines = text.split('\n')
    operation_started = False
    current_operation = None
    long_text_started = False
    
    for i, line in enumerate(lines):
        line = line.strip()
        if not line or line.startswith('_'):
            continue
            
        # Check if we've reached the operations section
        if "Oprn" in line and "Operation" in line:
            operation_started = True
            continue
            
        if operation_started:
            # Try to match operation row
            op_match = re.match(r'(\d{4})\s+([A-Z0-9-]+)\s+(\d+\.?\d*)\s+(\d+\.?\d*)\s+(\d+)\s+(\d+)\s+(\d+\.?\d*)\s*(\d*)', line)
            
            if op_match:
                if current_operation:
                    data["Operations"].append(current_operation)
                
                # Get the next line for additional plant info and operation
                next_line = lines[i + 1].strip() if i + 1 < len(lines) else ""
                next_next_line = lines[i + 2].strip() if i + 2 < len(lines) else ""
                
                # Extract plant number and operation description
                plant_number = ""
                operation_desc = ""
                
                if next_line:
                    # Check if next line contains a plant number
                    plant_match = re.match(r'^(\d+)\s*(.*)', next_line)
                    if plant_match:
                        plant_number = plant_match.group(1)
                        if plant_match.group(2):  # If there's text after the number
                            operation_desc = plant_match.group(2)
                        elif next_next_line and not next_next_line.startswith("Long Text"):
                            operation_desc = next_next_line
                    else:
                        operation_desc = next_line
                
                current_operation = {
                    "Oprn No": op_match.group(1),
                    "Wc/Plant": op_match.group(2),
                    "Plant Number": plant_number,
                    "Operation": operation_desc,
                    "Setup Time": op_match.group(3),
                    "Per Pc Time": op_match.group(4),
                    "Jmp Qty": op_match.group(5),
                    "Tot Qty": op_match.group(6),
                    "Allowed Time": op_match.group(7),
                    "Confirm No": op_match.group(8) if op_match.group(8) else "",
                    "Long Text": ""
                }
            elif current_operation:
                if "Long Text:" in line:
                    long_text_started = True
                    continue
                
                if long_text_started:
                    if current_operation["Long Text"]:
                        current_operation["Long Text"] += "\n" + line
                    else:
                        current_operation["Long Text"] = line
    
    # Add the last operation if exists
    if current_operation:
        data["Operations"].append(current_operation)
    
    # Extract document verification details from long text when operation is verification
    for operation in data["Operations"]:
        if "verification" in operation["Operation"].lower():
            doc_details = {}
            long_text = operation["Long Text"]
            
            # Extract document details using regex patterns
            doc_patterns = {
                "OARC Rev": r"OARC Rev\.\s*:\s*([^\n]+)",
                "Part Rev": r"Part Rev\.\s*:\s*([^\n]+)",
                "Drawing No": r"Drawing No\.\s*:\s*([^R]+)Rev\.\s*:\s*([^\n]+)",
                "Cad No": r"Cad No\.\s*:\s*([^R]+)Rev\.\s*:\s*([^\n]+)",
                "Stage Verification Doc": r"Stage Verification Document No\.\s*:\s*([^R]+)Rev\.\s*:\s*([^\n]+)",
                "Final Verification Doc": r"Final Verification Document No\.\s*:\s*([^R]+)Rev\.\s*:\s*([^\n]+)",
                "Raw Material Index Doc": r"Raw Material Index\s+Doc No\.\s*:\s*([\w\-]+)\s+Rev\.\s*:\s*(\d+)",
                "Plating Inspection Doc": r"Plating inspection Doc No\.\s*:([\w\-]+)\s+Rev\.\s*:\s*(\d+)",
                "MPP Doc": r"MPP Doc No\.\s*:\s*([^R]+)Rev\.\s*:\s*([^\n]+)"
            }
            
            for key, pattern in doc_patterns.items():
                match = re.search(pattern, long_text)
                if match:
                    if len(match.groups()) == 2:
                        doc_details[key] = {
                            "Number": match.group(1).strip(),
                            "Revision": match.group(2).strip()
                        }
                    else:
                        doc_details[key] = match.group(1).strip()
            
            data["Document Verification"] = doc_details
            break
    
    # Extract raw materials
    raw_materials_started = False
    raw_material_pattern = r'(\d{4})\s+(\w+)\s+([\w\s\-\.]+)\s+([\d\.]+)\s+(\w+)\s+([\d\.]+)'
    
    for i, line in enumerate(lines):
        line = line.strip()
        
        # Check if we've reached the raw materials section
        if "Item" in line and "Child Part No" in line:
            raw_materials_started = True
            continue
        
        if raw_materials_started and not line.startswith('_'):
            # Try to match raw material row
            raw_match = re.match(raw_material_pattern, line)
            if raw_match:
                raw_material = {
                    "Sl.No": raw_match.group(1),
                    "Child Part No": raw_match.group(2),
                    "Description": raw_match.group(3).strip(),
                    "Qty Per Set": raw_match.group(4),
                    "UoM": raw_match.group(5),
                    "Total Qty": raw_match.group(6)
                }
                data["Raw Materials"].append(raw_material)
        
        # End raw materials section if we hit another section
        if raw_materials_started and line.startswith('SPECIAL NOTE'):
            raw_materials_started = False
    
    return data

def show_manual_entry_forms(master_order_id):
    st.subheader("Manual Data Entry")
    
    # Work Center Entry
    st.write("Work Center Details")
    with st.form("work_center_form"):
        wc_code = st.text_input("Work Center Code")
        wc_description = st.text_input("Work Center Description")
        machine_name = st.text_input("Machine Name")
        machine_status = st.selectbox("Machine Status", 
                                    ["Active", "Maintenance", "Inactive"])
        
        if st.form_submit_button("Add Work Center"):
            try:
                with db_session:
                    # Create or get work center
                    work_center = WorkCenter.get(work_center_code=wc_code)
                    if not work_center:
                        work_center = WorkCenter(
                            work_center_code=wc_code,
                            description=wc_description
                        )
                    
                    # Add machine
                    WorkCenterMachine(
                        work_center=work_center,
                        machine_name=machine_name,
                        status=machine_status
                    )
                    st.success("Work Center and Machine added successfully!")
            except Exception as e:
                st.error(f"Error adding work center: {str(e)}")

    # Delivery Schedule Entry
    st.write("Delivery Schedule")
    with st.form("delivery_schedule_form"):
        scheduled_date = st.date_input("Scheduled Delivery Date")
        actual_date = st.date_input("Actual Delivery Date (if delivered)")
        status = st.selectbox("Delivery Status", 
                            ["Scheduled", "In Transit", "Delivered", "Delayed"])
        
        if st.form_submit_button("Add Delivery Schedule"):
            try:
                with db_session:
                    master_order = MasterOrder[master_order_id]
                    DeliverySchedule(
                        order=master_order,
                        scheduled_delivery_date=scheduled_date,
                        actual_delivery_date=actual_date if status == "Delivered" else None,
                        delivery_status=status
                    )
                    st.success("Delivery Schedule added successfully!")
            except Exception as e:
                st.error(f"Error adding delivery schedule: {str(e)}")

def main():
    st.title("OARC PDF Data Extractor")
    st.write("Upload your OARC PDF file to extract manufacturing routing plan details")
    
    uploaded_file = st.file_uploader("Choose a PDF file", type="pdf")
    
    if uploaded_file is not None:
        # Read PDF content
        pdf_content = io.BytesIO(uploaded_file.read())
        
        try:
            # Extract data
            data = extract_oarc_details(pdf_content)
            
            # Display header information
            st.subheader("Document Details")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.write(f"**Project Name:** {data['Project Name']}")
                st.write(f"**Sale Order:** {data['Sale Order']}")
                st.write(f"**Part No:** {data['Part No']}")
                st.write(f"**Plant:** {data['Plant']}")
            
            with col2:
                st.write(f"**Part Description:** {data['Part Desc']}")
                st.write(f"**Required Quantity:** {data['Required Qty']}")
                st.write(f"**Launched Quantity:** {data['Launched Qty']}")
                st.write(f"**WBS:** {data['WBS']}")
            
            with col3:
                st.write(f"**Rtg Seq No:** {data['Rtg Seq No']}")
                st.write(f"**Sequence No:** {data['Sequence No']}")
                st.write(f"**Prod Order No:** {data['Prod Order No']}")
            
            # Display operations table
            st.subheader("Operations")
            if data["Operations"]:
                df = pd.DataFrame(data["Operations"])
                column_order = ["Oprn No", "Wc/Plant", "Plant Number", "Operation", "Setup Time", 
                              "Per Pc Time", "Jmp Qty", "Tot Qty", "Allowed Time", 
                              "Confirm No", "Long Text"]
                df = df[column_order]
                st.dataframe(df, use_container_width=True)
            
            # Display raw materials table
            if data["Raw Materials"]:
                st.subheader("Raw Materials Details")
                raw_df = pd.DataFrame(data["Raw Materials"])
                column_order = ["Sl.No", "Child Part No", "Description", "Qty Per Set", "UoM", "Total Qty"]
                raw_df = raw_df[column_order]
                st.dataframe(raw_df, use_container_width=True)
            
            # Display document verification details if available
            if data["Document Verification"]:
                st.subheader("Document Verification Details")
                doc_data = []
                for key, value in data["Document Verification"].items():
                    if isinstance(value, dict):
                        doc_data.append({
                            "Document Type": key,
                            "Number": value["Number"],
                            "Revision": value["Revision"]
                        })
                    else:
                        doc_data.append({
                            "Document Type": key,
                            "Details": value
                        })
                
                doc_df = pd.DataFrame(doc_data)
                st.dataframe(doc_df, use_container_width=True)
                
            # Create Excel file and download button
            if data["Operations"]:
                document_details = {k: v for k, v in data.items() 
                          if k not in ["Operations", "Document Verification", "Raw Materials"]}
                
                # Create DataFrame for raw materials if they exist
                raw_materials_df = None
                if data["Raw Materials"]:
                    raw_materials_df = pd.DataFrame(data["Raw Materials"])
                    raw_materials_df = raw_materials_df[["Sl.No", "Child Part No", "Description", 
                                                       "Qty Per Set", "UoM", "Total Qty"]]
                
                # Create Excel file with all data
                excel_file = create_excel_file(document_details, df, raw_materials_df)
                
                with open(excel_file, "rb") as f:
                    excel_data = f.read()
                st.download_button(
                    label="Download Excel Report",
                    data=excel_data,
                    file_name="OARC_Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.warning("No operations data found in the PDF")
            
            # After successful extraction and before Excel creation
            with db_session:
                try:
                    # Create master order
                    master_order = MasterOrder(
                        project_name=data["Project Name"],
                        sale_order=data["Sale Order"],
                        part_number=data["Part No"],
                        wbs=data["WBS"],
                        part_description=data["Part Desc"],
                        total_operations=len(data["Operations"]),
                        plant=data["Plant"],
                        routing_sequence_no=int(data["Rtg Seq No"]),
                        required_quantity=int(data["Required Qty"]),
                        launched_quantity=int(data["Launched Qty"]),
                        production_order_no=data["Prod Order No"]
                    )
                    
                    # Insert Document References
                    if data["Document Verification"]:
                        for doc_type, doc_info in data["Document Verification"].items():
                            if isinstance(doc_info, dict):
                                DocumentReference(
                                    order=master_order,
                                    document_type=doc_type,
                                    document_number=doc_info["Number"],
                                    revision=doc_info["Revision"]
                                )
                            else:
                                DocumentReference(
                                    order=master_order,
                                    document_type=doc_type,
                                    document_number=doc_info,
                                    revision='--'
                                )

                    # Insert Raw Materials
                    if data["Raw Materials"]:
                        for raw_mat in data["Raw Materials"]:
                            RawMaterial(
                                order=master_order,
                                sl_no=raw_mat["Sl.No"],
                                child_part_no=raw_mat["Child Part No"],
                                description=raw_mat["Description"],
                                qty_per_set=float(raw_mat["Qty Per Set"]),
                                uom=raw_mat["UoM"],
                                total_qty=float(raw_mat["Total Qty"]),
                                is_available=True
                            )

                    # Insert Operations
                    if data["Operations"]:
                        for op in data["Operations"]:
                            # Create or get work center
                            work_center = WorkCenter.get(work_center_code=op["Wc/Plant"])
                            if not work_center:
                                work_center = WorkCenter(
                                    work_center_code=op["Wc/Plant"],
                                    description=op["Operation"]
                                )

                            Operation(
                                order=master_order,
                                work_center=work_center,
                                operation_number=int(op["Oprn No"]),
                                operation_description=op["Operation"],
                                setup_time=float(op["Setup Time"]),
                                per_piece_time=float(op["Per Pc Time"]),
                                jump_quantity=int(op["Jmp Qty"]),
                                total_quantity=int(op["Tot Qty"]),
                                allowed_time=float(op["Allowed Time"]),
                                actual_time=0.0,  # Default value
                                confirmation_number=op["Confirm No"] if op["Confirm No"] else None
                            )

                    # Show manual entry forms after successful database insertion
                    show_manual_entry_forms(master_order.order_id)

                    st.success("All data successfully saved to database!")
                    
                except Exception as e:
                    st.error(f"Error saving to database: {str(e)}")
                    raise e
            
        except Exception as e:
            st.error(f"Error processing PDF: {str(e)}")

if __name__ == "__main__":
    # Initialize database only if not already initialized
    if not db.provider:
        init_database()
    main()

     # TO RUN
# uvicorn app.main:app --host 172.18.101.47 --port 4567 --reload  